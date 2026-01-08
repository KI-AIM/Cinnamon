import argparse
import csv
import statistics
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import matplotlib.pyplot as plt
except ImportError:  # pragma: no cover - optional dependency
    plt = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None

DEFAULT_DATASET = "cardiovascular"
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_ROOT = BASE_DIR / "experiment_results" / DEFAULT_DATASET

SUMMARY_XLSX = "evaluation_summary.xlsx"
SUMMARY_CSV = "evaluation_summary.csv"

LABELS = ["real", "anon", "anon_synth"]
COLORS = {
    "real": "#333333",
    "anon": "#1f77b4",
    "anon_synth": "#2ca02c",
}


def read_summary(summary_path: Path) -> List[Dict[str, Any]]:
    if summary_path.suffix.lower() == ".xlsx":
        return read_summary_xlsx(summary_path)
    return read_summary_csv(summary_path)


def read_summary_xlsx(summary_path: Path) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read Excel files. Install it with: pip install openpyxl")
    workbook = load_workbook(summary_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows: List[Dict[str, Any]] = []
    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data_rows.append(row_dict)
    return data_rows


def read_summary_csv(summary_path: Path) -> List[Dict[str, Any]]:
    data_rows: List[Dict[str, Any]] = []
    with summary_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            data_rows.append(row)
    return data_rows


def as_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value))
    except ValueError:
        return None


def summary_value(row: Dict[str, Any], key: str) -> Optional[float]:
    return as_float(row.get(key))


def get_metric(row: Dict[str, Any], base_key: str, label: str, fallback_key: Optional[str] = None) -> Optional[float]:
    value = summary_value(row, f"{base_key}_{label}")
    if value is None and fallback_key:
        value = summary_value(row, f"{fallback_key}_{label}")
    return value


def compute_stats(values: List[float]) -> Tuple[Optional[float], Optional[float]]:
    if not values:
        return None, None
    if len(values) == 1:
        return values[0], 0.0
    return statistics.fmean(values), statistics.stdev(values)


def collect_metric(rows: List[Dict[str, Any]], metric: str, label: str) -> List[float]:
    values: List[float] = []
    for row in rows:
        value = summary_value(row, f"{metric}_{label}")
        if value is not None:
            values.append(value)
    return values


def plot_privacy_utility_tradeoff(rows: List[Dict[str, Any]], output_path: Path) -> None:
    if plt is None:
        raise RuntimeError("matplotlib is required to plot figures. Install it with: pip install matplotlib")

    fig, ax = plt.subplots(figsize=(9, 6))
    label_used = set()

    for row in rows:
        points = {}
        for label in LABELS:
            risk = get_metric(row, "privacy_score", label)
            utility = get_metric(row, "combined_utility", label, fallback_key="overall_utility")
            if risk is None or utility is None:
                continue
            points[label] = (risk, utility)

        if "real" in points and "anon" in points:
            ax.annotate(
                "",
                xy=points["anon"],
                xytext=points["real"],
                arrowprops=dict(arrowstyle="->", color="#888888", alpha=0.4),
            )
        if "anon" in points and "anon_synth" in points:
            ax.annotate(
                "",
                xy=points["anon_synth"],
                xytext=points["anon"],
                arrowprops=dict(arrowstyle="->", color="#888888", alpha=0.4),
            )

        for label, coords in points.items():
            plot_label = label if label not in label_used else None
            ax.scatter(
                coords[0],
                coords[1],
                color=COLORS[label],
                s=40,
                label=plot_label,
                alpha=0.8,
            )
            label_used.add(label)

    ax.set_title("Privacy-Utility Tradeoff")
    ax.set_xlabel("Privacy Score (1 - risk)")
    ax.set_ylabel("Combined Utility")
    ax.set_xlim(0.0, 1.0)
    ax.set_ylim(0.0, 1.0)
    ax.grid(True, linestyle="--", alpha=0.3)
    ax.legend(frameon=False)
    fig.tight_layout()
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def plot_summary_bars(rows: List[Dict[str, Any]], output_path: Path) -> None:
    if plt is None:
        raise RuntimeError("matplotlib is required to plot figures. Install it with: pip install matplotlib")

    metrics = [
        ("combined_utility", "Combined Utility"),
        ("privacy_score", "Privacy Score (1 - risk)"),
    ]

    fig, axes = plt.subplots(1, 2, figsize=(11, 4.5))
    for ax, (metric, title) in zip(axes, metrics):
        means = []
        stds = []
        for label in LABELS:
            values = collect_metric(rows, metric, label)
            mean, std = compute_stats(values)
            means.append(mean)
            stds.append(std)
        ax.bar(LABELS, means, yerr=stds, color=[COLORS[l] for l in LABELS], alpha=0.85)
        ax.set_title(title)
        ax.set_xticklabels(LABELS, rotation=0)
        ax.grid(True, axis="y", linestyle="--", alpha=0.3)

    fig.tight_layout()
    fig.savefig(output_path, dpi=200)
    plt.close(fig)


def resolve_summary_path(root: Path, summary_path: Optional[str]) -> Path:
    if summary_path:
        return Path(summary_path)
    xlsx_path = root / SUMMARY_XLSX
    if xlsx_path.exists():
        return xlsx_path
    csv_path = root / SUMMARY_CSV
    if csv_path.exists():
        return csv_path
    raise FileNotFoundError(
        f"No summary file found. Expected {xlsx_path} or {csv_path}."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Visualize evaluation summary results.")
    parser.add_argument(
        "--root",
        default=str(DEFAULT_ROOT),
        help="Base folder containing evaluation_summary.xlsx or .csv.",
    )
    parser.add_argument(
        "--summary",
        help="Explicit path to the summary file (xlsx or csv).",
    )
    args = parser.parse_args()

    root = Path(args.root)
    summary_path = resolve_summary_path(root, args.summary)
    rows = read_summary(summary_path)
    if not rows:
        raise RuntimeError(f"No data found in {summary_path}")

    output_dir = summary_path.parent
    plot_privacy_utility_tradeoff(rows, output_dir / "privacy_utility_tradeoff.png")
    plot_summary_bars(rows, output_dir / "summary_bars.png")
    print(f"Saved figures to {output_dir}")


if __name__ == "__main__":
    main()
