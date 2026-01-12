import argparse
import csv
import json
import math
import re
import statistics
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D
    from matplotlib.colors import ListedColormap, Normalize
except ImportError:  # pragma: no cover - optional dependency
    plt = None
    ListedColormap = None
    Normalize = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None

try:
    import numpy as np
except ImportError:  # pragma: no cover - optional dependency
    np = None

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_ROOT = BASE_DIR / "experiment_results"

SUMMARY_XLSX = "evaluation_summary.xlsx"
SUMMARY_CSV = "evaluation_summary.csv"

LABELS = ["real", "anon", "anon_synth"]
LABEL_TITLES = {
    "real": "Real",
    "anon": "Anonymized",
    "anon_synth": "Anon+Synth",
}
COLORS = {
    "real": "#333333",
    "anon": "#1b6ca8",
    "anon_synth": "#2a9d8f",
}
MARKERS = {
    "real": "o",
    "anon": "o",
    "anon_synth": "o",
}


def read_summary(summary_path: Path) -> List[Dict[str, Any]]:
    if summary_path.suffix.lower() == ".xlsx":
        return read_summary_xlsx(summary_path)
    return read_summary_csv(summary_path)


def read_summary_xlsx(summary_path: Path) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read Excel files. Install it with: pip install openpyxl")
    workbook = load_workbook(summary_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows: List[Dict[str, Any]] = []
    for row in rows[1:]:
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        data_rows.append(row_dict)
    return data_rows


def read_summary_csv(summary_path: Path) -> List[Dict[str, Any]]:
    data_rows: List[Dict[str, Any]] = []
    with summary_path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            data_rows.append(row)
    return data_rows


def as_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value))
    except ValueError:
        return None


def summary_value(row: Dict[str, Any], key: str) -> Optional[float]:
    return as_float(row.get(key))


def get_metric(row: Dict[str, Any], base_key: str, label: str, fallback_key: Optional[str] = None) -> Optional[float]:
    value = summary_value(row, f"{base_key}_{label}")
    if value is None and fallback_key:
        value = summary_value(row, f"{fallback_key}_{label}")
    return value


def compute_stats(values: List[float]) -> Tuple[Optional[float], Optional[float]]:
    if not values:
        return None, None
    if len(values) == 1:
        return values[0], 0.0
    return statistics.fmean(values), statistics.stdev(values)


def collect_metric(
    rows: List[Dict[str, Any]],
    metric: str,
    label: str,
    fallback_key: Optional[str] = None,
) -> List[float]:
    values: List[float] = []
    for row in rows:
        value = get_metric(row, metric, label, fallback_key=fallback_key)
        if value is not None:
            values.append(value)
    return values


def prettify_dataset_name(dataset_name: str) -> str:
    return re.sub(r"[_-]+", " ", dataset_name).strip().title()


def flatten_axes(axes):
    if isinstance(axes, list):
        return axes
    if hasattr(axes, "shape"):
        if len(axes.shape) == 0:
            return [axes]
        if len(axes.shape) == 1:
            return list(axes)
        return [ax for row in axes for ax in row]
    return [axes]


def format_dataset_title(dataset_name: str, size: Optional[int]) -> str:
    base = prettify_dataset_name(dataset_name)
    if size:
        return f"{base} (N={size:,})"
    return base


def resolve_dataset_csv(dataset_name: str) -> Optional[Path]:
    dataset_dir = BASE_DIR / "dataset_configs" / dataset_name
    if not dataset_dir.exists():
        return None
    preferred = dataset_dir / "original-dataset.csv"
    if preferred.exists():
        return preferred
    csv_files = sorted(dataset_dir.glob("*.csv"))
    if csv_files:
        return csv_files[0]
    return None


def get_dataset_size(dataset_name: str) -> Optional[int]:
    dataset_path = resolve_dataset_csv(dataset_name)
    if dataset_path is None:
        return None
    try:
        with dataset_path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.reader(handle)
            rows = 0
            for _ in reader:
                rows += 1
        if rows <= 1:
            return 0
        return rows - 1
    except OSError:
        return None


def plot_privacy_utility_tradeoff(
    dataset_rows: Dict[str, List[Dict[str, Any]]],
    dataset_sizes: Dict[str, Optional[int]],
    output_path: Path,
) -> None:
    if plt is None:
        raise RuntimeError("matplotlib is required to plot figures. Install it with: pip install matplotlib")

    plt.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.sans-serif": ["DejaVu Sans", "Arial", "Liberation Sans"],
            "axes.titlesize": 11,
            "axes.labelsize": 9,
            "xtick.labelsize": 8,
            "ytick.labelsize": 8,
            "axes.facecolor": "#fbfbfb",
            "figure.facecolor": "white",
            "grid.color": "#c9c9c9",
            "grid.alpha": 0.35,
            "grid.linewidth": 0.6,
        }
    )

    dataset_names = sorted(dataset_rows.keys())
    if not dataset_names:
        raise RuntimeError("No datasets found to plot.")

    ncols = 3 if len(dataset_names) > 3 else len(dataset_names)
    nrows = math.ceil(len(dataset_names) / ncols)
    fig, axes = plt.subplots(
        nrows=nrows,
        ncols=ncols,
        figsize=(4.6 * ncols, 3.6 * nrows),
        sharex=True,
        sharey=True,
    )
    axes_list = flatten_axes(axes)

    for ax, dataset_name in zip(axes_list, dataset_names):
        rows = dataset_rows[dataset_name]
        run_points: Dict[str, List[Tuple[float, float]]] = {label: [] for label in LABELS}
        mean_points: Dict[str, Tuple[float, float]] = {}

        for row in rows:
            for label in LABELS:
                risk = get_metric(row, "privacy_score", label)
                utility = get_metric(row, "combined_utility", label, fallback_key="overall_utility")
                if risk is None or utility is None:
                    continue
                run_points[label].append((risk, utility))

        for label in LABELS:
            points = run_points[label]
            if points:
                xs, ys = zip(*points)
                ax.scatter(
                    xs,
                    ys,
                    s=14,
                    color=COLORS[label],
                    alpha=0.22,
                    edgecolor="none",
                    zorder=1,
                )

            risks = collect_metric(rows, "privacy_score", label)
            utilities = collect_metric(rows, "combined_utility", label, fallback_key="overall_utility")
            mean_risk, std_risk = compute_stats(risks)
            mean_utility, std_utility = compute_stats(utilities)

            if mean_risk is None or mean_utility is None:
                continue

            mean_points[label] = (mean_risk, mean_utility)
            ax.scatter(
                mean_risk,
                mean_utility,
                s=46,
                color=COLORS[label],
                marker=MARKERS[label],
                zorder=5,
            )

        if "real" in mean_points and "anon" in mean_points:
            ax.annotate(
                "",
                xy=mean_points["anon"],
                xytext=mean_points["real"],
                arrowprops=dict(arrowstyle="->", color="#555555", lw=1.0, alpha=0.7),
                zorder=3,
            )
        if "anon" in mean_points and "anon_synth" in mean_points:
            ax.annotate(
                "",
                xy=mean_points["anon_synth"],
                xytext=mean_points["anon"],
                arrowprops=dict(arrowstyle="->", color="#555555", lw=1.0, alpha=0.7),
                zorder=3,
            )

        ax.set_title(format_dataset_title(dataset_name, dataset_sizes.get(dataset_name)))
        ax.set_xlim(0.0, 1.0)
        ax.set_ylim(0.0, 1.0)
        ax.grid(True, which="major", linestyle="-", alpha=0.35)
        ax.minorticks_on()
        ax.grid(True, which="minor", linestyle=":", alpha=0.18)
        ax.tick_params(direction="out", length=3, width=0.8, colors="#444444")
        ax.spines["top"].set_visible(True)
        ax.spines["right"].set_visible(True)
        ax.spines["top"].set_color("#111111")
        ax.spines["right"].set_color("#111111")
        ax.spines["left"].set_color("#666666")
        ax.spines["bottom"].set_color("#666666")

    for ax in axes_list[len(dataset_names) :]:
        ax.axis("off")

    fig.supxlabel("Privacy Score", y=0.015)
    fig.supylabel("Utility Score")

    legend_handles = [
        Line2D(
            [0],
            [0],
            marker=MARKERS[label],
            color="none",
            markerfacecolor=COLORS[label],
            markeredgecolor="white",
            markeredgewidth=0.8,
            markersize=8,
            label=LABEL_TITLES[label],
        )
        for label in LABELS
    ]
    fig.legend(
        handles=legend_handles,
        loc="upper center",
        ncol=len(LABELS),
        frameon=False,
        bbox_to_anchor=(0.5, 0.98),
    )
    fig.tight_layout(rect=(0, 0.02, 1, 0.93))
    fig.savefig(output_path, dpi=300, bbox_inches="tight", pad_inches=0.03)
    plt.close(fig)


def plot_privacy_utility_heatmap(
    dataset_rows: Dict[str, List[Dict[str, Any]]],
    output_path: Path,
) -> None:
    if plt is None or ListedColormap is None or Normalize is None:
        raise RuntimeError("matplotlib is required to plot figures. Install it with: pip install matplotlib")

    dataset_names = sorted(dataset_rows.keys())
    if not dataset_names:
        raise RuntimeError("No datasets found to plot.")

    labels = ["real", "anon", "anon_synth"]
    columns = [
        ("Utility", "combined_utility", "overall_utility"),
        ("Privacy", "privacy_score", None),
    ]
    gap_width = 0.5
    col_specs: List[Dict[str, Any]] = []
    for label_index, label in enumerate(labels):
        for metric_title, metric_key, fallback in columns:
            label_title = LABEL_TITLES[label]
            col_specs.append(
                {
                    "label": f"{metric_title}\n{label_title}",
                    "metric_title": metric_title,
                    "metric_key": metric_key,
                    "fallback": fallback,
                    "label_key": label,
                    "is_gap": False,
                }
            )
        if label_index < len(labels) - 1:
            col_specs.append({"label": "", "metric_key": None, "fallback": None, "is_gap": True})

    data: List[List[float]] = []
    stds: List[List[Optional[float]]] = []

    for dataset_name in dataset_names:
        rows = dataset_rows[dataset_name]
        row_values: List[float] = []
        row_stds: List[Optional[float]] = []
        for spec in col_specs:
            if spec["is_gap"]:
                row_values.append(float("nan"))
                row_stds.append(None)
                continue
            values = collect_metric(
                rows,
                spec["metric_key"],
                spec["label_key"],
                fallback_key=spec["fallback"],
            )
            mean, std = compute_stats(values)
            row_values.append(mean if mean is not None else float("nan"))
            row_stds.append(std if mean is not None else None)
        data.append(row_values)
        stds.append(row_stds)

    csv_output = output_path.with_suffix(".csv")
    json_output = output_path.with_suffix(".json")
    header = ["dataset"]
    data_indices: List[int] = []
    for idx, spec in enumerate(col_specs):
        if spec["is_gap"]:
            continue
        data_indices.append(idx)
        base = f"{spec['metric_title'].lower()}_{spec['label_key']}"
        header.extend([f"{base}_mean", f"{base}_std"])
    with csv_output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for dataset_name, row_values, row_stds in zip(dataset_names, data, stds):
            row = [dataset_name]
            for idx in data_indices:
                value = row_values[idx]
                std = row_stds[idx]
                row.append("" if value != value else value)
                row.append("" if std is None else std)
            writer.writerow(row)

    json_rows: List[Dict[str, Any]] = []
    for dataset_name, row_values, row_stds in zip(dataset_names, data, stds):
        payload: Dict[str, Any] = {"dataset": dataset_name}
        for idx, spec in enumerate(col_specs):
            if spec["is_gap"]:
                continue
            base = f"{spec['metric_title'].lower()}_{spec['label_key']}"
            value = row_values[idx]
            std = row_stds[idx]
            payload[f"{base}_mean"] = None if value != value else value
            payload[f"{base}_std"] = None if std is None else std
        json_rows.append(payload)
    with json_output.open("w", encoding="utf-8") as handle:
        json.dump(json_rows, handle, indent=2)

    if ListedColormap is None or np is None:
        raise RuntimeError("numpy and matplotlib are required to plot heatmaps. Install them with: pip install numpy")

    base = plt.get_cmap("YlGnBu")
    colors = base(np.linspace(0.82, 0, 256))
    cmap = ListedColormap(colors)
    cmap.set_bad("#ffffff")
    norm = Normalize(vmin=0, vmax=1)

    fig_height = max(2.5, 0.55 * len(dataset_names) + 1.5)
    fig, ax = plt.subplots(figsize=(10.2, fig_height))

    col_widths = [gap_width if spec["is_gap"] else 1.0 for spec in col_specs]
    x_edges = [0.0]
    for width in col_widths:
        x_edges.append(x_edges[-1] + width)
    y_edges = list(range(len(dataset_names) + 1))

    im = ax.pcolormesh(x_edges, y_edges, data, cmap=cmap, norm=norm, shading="flat")
    ax.invert_yaxis()

    x_tick_positions = []
    x_tick_labels = []
    centers = []
    for spec, width, start in zip(col_specs, col_widths, x_edges[:-1]):
        center = start + width / 2
        centers.append(center)
        if not spec["is_gap"]:
            x_tick_positions.append(center)
            x_tick_labels.append(spec["metric_title"])

    ax.set_xticks(x_tick_positions)
    ax.set_xticklabels(x_tick_labels)
    ax.set_yticks([y + 0.5 for y in range(len(dataset_names))])
    ax.set_yticklabels([prettify_dataset_name(name) for name in dataset_names])
    ax.tick_params(axis="x", labelsize=9, pad=6)
    ax.tick_params(axis="y", labelsize=9)

    ax.set_xticks(x_edges, minor=True)
    ax.set_yticks(y_edges, minor=True)
    ax.grid(which="minor", color="#e1e1e1", linewidth=0.6)
    ax.tick_params(which="minor", bottom=False, left=False)

    group_centers = []
    col_index = 0
    for label_index, label in enumerate(labels):
        first = col_index
        last = col_index + len(columns) - 1
        group_center = (x_edges[first] + x_edges[last + 1]) / 2
        group_centers.append((group_center, LABEL_TITLES[label]))
        col_index += len(columns)
        if label_index < len(labels) - 1:
            col_index += 1

    for center, title in group_centers:
            ax.text(
                center,
                1.03,
                title,
                transform=ax.get_xaxis_transform(),
                ha="center",
                va="bottom",
                fontsize=9,
                fontweight="normal",
                color="#222222",
            )

    for i in range(len(dataset_names)):
        for j, center in enumerate(centers):
            value = data[i][j]
            std = stds[i][j]
            if value != value:
                text = ""
                text_color = "#666666"
            else:
                if std is None:
                    text = f"{value:.2f}"
                else:
                    text = f"{value:.2f}+/-{std:.2f}"
                text_color = "#111111"
            ax.text(
                center,
                i + 0.5,
                text,
                ha="center",
                va="center",
                fontsize=9,
                fontweight="normal",
                color=text_color,
            )

    cbar = fig.colorbar(im, ax=ax, fraction=0.035, pad=0.02)
    cbar.ax.set_ylabel("Score", rotation=90, labelpad=8)
    cbar.ax.tick_params(labelsize=8)

    ax.set_xlabel("")
    ax.set_ylabel("")
    ax.spines["top"].set_visible(False)
    fig.tight_layout(rect=(0, 0, 1, 0.94))
    fig.savefig(output_path, dpi=300, bbox_inches="tight", pad_inches=0.03)
    plt.close(fig)


def resolve_summary_path(root: Path, summary_path: Optional[str]) -> Path:
    if summary_path:
        return Path(summary_path)
    xlsx_path = root / SUMMARY_XLSX
    if xlsx_path.exists():
        return xlsx_path
    csv_path = root / SUMMARY_CSV
    if csv_path.exists():
        return csv_path
    raise FileNotFoundError(f"No summary file found. Expected {xlsx_path} or {csv_path}.")


def collect_dataset_summaries(root: Path, summary_path: Optional[str]) -> Dict[str, Path]:
    if summary_path:
        summary = Path(summary_path)
        return {summary.parent.name: summary}

    if (root / SUMMARY_XLSX).exists() or (root / SUMMARY_CSV).exists():
        return {root.name: resolve_summary_path(root, None)}

    datasets: Dict[str, Path] = {}
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        try:
            datasets[child.name] = resolve_summary_path(child, None)
        except FileNotFoundError:
            continue
    return datasets


def main() -> None:
    parser = argparse.ArgumentParser(description="Visualize evaluation summary results.")
    parser.add_argument(
        "--root",
        default=str(DEFAULT_ROOT),
        help="Base folder containing dataset subfolders with evaluation_summary.xlsx or .csv.",
    )
    parser.add_argument(
        "--summary",
        help="Explicit path to the summary file (xlsx or csv).",
    )
    args = parser.parse_args()

    root = Path(args.root)
    if not root.exists():
        raise FileNotFoundError(f"Root folder not found: {root}")

    dataset_summaries = collect_dataset_summaries(root, args.summary)
    if not dataset_summaries:
        raise RuntimeError(f"No evaluation summaries found under {root}")

    dataset_rows: Dict[str, List[Dict[str, Any]]] = {}
    dataset_sizes: Dict[str, Optional[int]] = {}
    for dataset_name, summary_path in dataset_summaries.items():
        rows = read_summary(summary_path)
        if not rows:
            continue
        dataset_rows[dataset_name] = rows
        dataset_sizes[dataset_name] = get_dataset_size(dataset_name)

    if not dataset_rows:
        raise RuntimeError("No data found in any evaluation summary.")

    output_dir = root
    plot_privacy_utility_tradeoff(
        dataset_rows,
        dataset_sizes,
        output_dir / "privacy_utility_tradeoff_overview.png",
    )
    plot_privacy_utility_heatmap(
        dataset_rows,
        output_dir / "privacy_utility_heatmap_overview.png",
    )
    print(f"Saved figures to {output_dir}")


if __name__ == "__main__":
    main()
